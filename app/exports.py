from flask import Response
from flask_appbuilder import BaseView, expose
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import date

from .extensions import appbuilder
from .models import Reserva, Pago, ReservaServicio, Cliente


# ── Helpers PDF ──────────────────────────────────────────────────────────────

HEADER_COLOR  = colors.HexColor("#3c3c3c")
ROW_ODD       = colors.HexColor("#f9f9f9")
ROW_EVEN      = colors.white
ACCENT_COLOR  = colors.HexColor("#4e73df")

def base_table_style(n_cols):
    return TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0),  HEADER_COLOR),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0),  9),
        ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE",     (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [ROW_ODD, ROW_EVEN]),
        ("GRID",         (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("TOPPADDING",   (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
    ])

def build_pdf(title, headers, rows, landscape_mode=False):
    buffer = BytesIO()
    pagesize = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(buffer, pagesize=pagesize,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 textColor=ACCENT_COLOR, fontSize=16, spaceAfter=6)
    date_style  = ParagraphStyle("date", parent=styles["Normal"],
                                 fontSize=8, textColor=colors.grey, spaceAfter=12)

    col_width = (pagesize[0] - 3*cm) / len(headers)
    table_data = [headers] + rows
    tbl = Table(table_data, colWidths=[col_width] * len(headers), repeatRows=1)
    tbl.setStyle(base_table_style(len(headers)))

    story = [
        Paragraph(title, title_style),
        Paragraph(f"Generado el {date.today().strftime('%d/%m/%Y')}", date_style),
        tbl,
    ]
    doc.build(story)
    buffer.seek(0)
    return buffer


# ── Helpers Excel ─────────────────────────────────────────────────────────────

def base_excel(title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font    = Font(bold=True, color="FFFFFF", size=10)
    header_fill    = PatternFill("solid", fgColor="3c3c3c")
    header_align   = Alignment(horizontal="center", vertical="center")
    row_odd_fill   = PatternFill("solid", fgColor="F9F9F9")
    thin_border    = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=13, color="4e73df")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    # Cabeceras
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 20

    # Datos
    for r_idx, row in enumerate(rows, 3):
        fill = row_odd_fill if r_idx % 2 == 0 else PatternFill()
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Ancho automático (saltamos celdas mergeadas)
    for col in ws.columns:
        col_letter = None
        max_len = 10
        for c in col:
            try:
                if c.column_letter:
                    col_letter = c.column_letter
                val_len = len(str(c.value or ""))
                if val_len > max_len:
                    max_len = val_len
            except AttributeError:
                continue
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── Vistas de exportación ─────────────────────────────────────────────────────

class ExportReservasPDF(BaseView):
    route_base = "/exportar/reservas/pdf"

    @expose("/")
    def index(self):
        reservas = Reserva.query.all()
        headers = ["#", "Reservante", "Teléfono", "Habitación", "Tipo", "Entrada", "Salida", "Estado"]
        rows = [
            [
                r.id,
                r.nombre_reservante,
                r.telefono_contacto or "-",
                f"Hab. {r.habitacion.numero}",
                r.habitacion.tipo.nombre,
                str(r.fecha_entrada),
                str(r.fecha_salida),
                r.estado,
            ]
            for r in reservas
        ]
        pdf = build_pdf("Reporte de Reservas", headers, rows, landscape_mode=True)
        return Response(pdf, mimetype="application/pdf",
                        headers={"Content-Disposition": "attachment;filename=reservas.pdf"})


class ExportReservasExcel(BaseView):
    route_base = "/exportar/reservas/excel"

    @expose("/")
    def index(self):
        reservas = Reserva.query.all()
        headers = ["#", "Reservante", "Teléfono", "Habitación", "Tipo", "Entrada", "Salida", "Estado"]
        rows = [
            [
                r.id,
                r.nombre_reservante,
                r.telefono_contacto or "-",
                f"Hab. {r.habitacion.numero}",
                r.habitacion.tipo.nombre,
                str(r.fecha_entrada),
                str(r.fecha_salida),
                r.estado,
            ]
            for r in reservas
        ]
        excel = base_excel("Reporte de Reservas", headers, rows)
        return Response(excel,
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment;filename=reservas.xlsx"})


class ExportPagosPDF(BaseView):
    route_base = "/exportar/pagos/pdf"

    @expose("/")
    def index(self):
        pagos = Pago.query.order_by(Pago.fecha.desc(), Pago.id.desc()).all()
        total = sum(p.monto for p in pagos)
        headers = ["# Pago", "Referencia", "Origen", "Nombre", "Habitación", "Monto (Bs)", "Método", "Fecha"]
        rows = []
        for p in pagos:
            if p.reserva:
                referencia = p.reserva.id
                origen = "Reserva"
                nombre = p.reserva.nombre_reservante
                habitacion = f"Hab. {p.reserva.habitacion.numero}" if p.reserva.habitacion else "-"
            elif p.cliente:
                referencia = p.cliente.id
                origen = "Cliente"
                nombre = p.cliente.nombre_completo
                habitacion = f"Hab. {p.cliente.habitacion.numero}" if p.cliente.habitacion else "-"
            else:
                referencia = "-"
                origen = "-"
                nombre = "-"
                habitacion = "-"
            rows.append([
                p.id,
                referencia,
                origen,
                nombre,
                habitacion,
                f"{p.monto:.2f}",
                p.metodo,
                str(p.fecha),
            ])
        rows.append(["", "", "", "", "TOTAL", f"{total:.2f}", "", ""])
        pdf = build_pdf("Reporte de Pagos", headers, rows, landscape_mode=True)
        return Response(pdf, mimetype="application/pdf",
                        headers={"Content-Disposition": "attachment;filename=pagos.pdf"})


class ExportPagosExcel(BaseView):
    route_base = "/exportar/pagos/excel"

    @expose("/")
    def index(self):
        pagos = Pago.query.order_by(Pago.fecha.desc(), Pago.id.desc()).all()
        total = sum(p.monto for p in pagos)
        headers = ["# Pago", "Referencia", "Origen", "Nombre", "Habitación", "Monto (Bs)", "Método", "Fecha"]
        rows = []
        for p in pagos:
            if p.reserva:
                referencia = p.reserva.id
                origen = "Reserva"
                nombre = p.reserva.nombre_reservante
                habitacion = f"Hab. {p.reserva.habitacion.numero}" if p.reserva.habitacion else "-"
            elif p.cliente:
                referencia = p.cliente.id
                origen = "Cliente"
                nombre = p.cliente.nombre_completo
                habitacion = f"Hab. {p.cliente.habitacion.numero}" if p.cliente.habitacion else "-"
            else:
                referencia = "-"
                origen = "-"
                nombre = "-"
                habitacion = "-"
            rows.append([
                p.id,
                referencia,
                origen,
                nombre,
                habitacion,
                round(p.monto, 2),
                p.metodo,
                str(p.fecha),
            ])
        rows.append(["", "", "", "", "TOTAL", round(total, 2), "", ""])
        excel = base_excel("Reporte de Pagos", headers, rows)
        return Response(excel,
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment;filename=pagos.xlsx"})


class ExportClientesPDF(BaseView):
    route_base = "/exportar/clientes/pdf"

    @expose("/")
    def index(self):
        clientes = Cliente.query.all()
        headers = ["Hora", "Nombre", "C.I.", "Nacionalidad", "Procedencia", "Teléfono", "Habitación"]
        rows = [
            [
                c.hora_ingreso or "-",
                c.nombre_completo,
                c.ci,
                c.nacionalidad or "-",
                c.procedencia or "-",
                c.telefono or "-",
                f"Hab. {c.habitacion.numero}" if c.habitacion else "-",
            ]
            for c in clientes
        ]
        pdf = build_pdf("Registro de Clientes", headers, rows, landscape_mode=True)
        return Response(pdf, mimetype="application/pdf",
                        headers={"Content-Disposition": "attachment;filename=clientes.pdf"})


class ExportClientesExcel(BaseView):
    route_base = "/exportar/clientes/excel"

    @expose("/")
    def index(self):
        clientes = Cliente.query.all()
        headers = ["Hora", "Nombre", "C.I.", "Nacionalidad", "Procedencia", "Profesión", "Estado Civil", "Teléfono", "Habitación"]
        rows = [
            [
                c.hora_ingreso or "-",
                c.nombre_completo,
                c.ci,
                c.nacionalidad or "-",
                c.procedencia or "-",
                c.profesion or "-",
                c.estado_civil or "-",
                c.telefono or "-",
                f"Hab. {c.habitacion.numero}" if c.habitacion else "-",
            ]
            for c in clientes
        ]
        excel = base_excel("Registro de Clientes", headers, rows)
        return Response(excel,
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment;filename=clientes.xlsx"})


# ── Registro ──────────────────────────────────────────────────────────────────
def register_exports():
    appbuilder.add_view_no_menu(ExportReservasPDF)
    appbuilder.add_view_no_menu(ExportReservasExcel)
    appbuilder.add_view_no_menu(ExportPagosPDF)
    appbuilder.add_view_no_menu(ExportPagosExcel)
    appbuilder.add_view_no_menu(ExportClientesPDF)
    appbuilder.add_view_no_menu(ExportClientesExcel)
